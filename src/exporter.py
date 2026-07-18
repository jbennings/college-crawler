from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_ROOT / "data" / "output"


def get_timestamp() -> str:
    """
    Returns a timestamp for export filenames.
    """
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def ensure_output_dir():
    """
    Makes sure the output directory exists.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def format_role_target_excel(excel_path: Path) -> None:
    """
    Applies review-friendly formatting to the role-target Excel export.
    """
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    worksheet.title = "Role Target Contacts"

    # Freeze the header row and enable filtering.
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    # Header styling.
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1D3557",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    thin_gray_border = Border(
        bottom=Side(
            style="thin",
            color="B7B7B7",
        )
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_gray_border

    worksheet.row_dimensions[1].height = 34

    # Column widths by exported field.
    column_widths = {
        "A": 30,  # organization
        "B": 24,  # name
        "C": 38,  # title
        "D": 30,  # department
        "E": 18,  # phone
        "F": 32,  # email
        "G": 26,  # role_categories
        "H": 30,  # matched_role_terms
        "I": 18,  # role_priority_score
        "J": 20,  # source_type
        "K": 35,  # source_page_title
        "L": 52,  # source_url
        "M": 25,  # additional_source_types
        "N": 38,  # additional_source_page_titles
        "O": 52,  # additional_source_urls
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    # Body styling.
    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        row[8].alignment = Alignment(
            horizontal="center",
            vertical="top",
        )

    # Keep phone numbers and email addresses as text.
    for row_number in range(2, worksheet.max_row + 1):
        worksheet[f"E{row_number}"].number_format = "@"
        worksheet[f"F{row_number}"].number_format = "@"

    # Highlight highest-priority contacts.
    high_priority_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )
    high_priority_font = Font(
        bold=True,
        color="7F6000",
    )

    worksheet.conditional_formatting.add(
        f"I2:I{worksheet.max_row}",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["100"],
            fill=high_priority_fill,
            font=high_priority_font,
        ),
    )

    # Add subtle alternating row shading for easier review.
    alternate_fill = PatternFill(
        fill_type="solid",
        fgColor="F4F6F8",
    )

    for row_number in range(2, worksheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in worksheet[row_number]:
                cell.fill = alternate_fill

    # Restore high-priority row emphasis across the full row.
    for row_number in range(2, worksheet.max_row + 1):
        priority_value = worksheet[f"I{row_number}"].value

        try:
            priority_score = int(priority_value)
        except (TypeError, ValueError):
            priority_score = 0

        if priority_score >= 100:
            for cell in worksheet[row_number]:
                cell.fill = high_priority_fill

            worksheet[f"B{row_number}"].font = Font(
                bold=True,
                color="7F6000",
            )
            worksheet[f"C{row_number}"].font = Font(
                bold=True,
                color="7F6000",
            )

    # Make URLs visually recognizable.
    url_columns = ["L", "O"]

    for column_letter in url_columns:
        for row_number in range(2, worksheet.max_row + 1):
            cell = worksheet[f"{column_letter}{row_number}"]

            if cell.value:
                cell.font = Font(
                    color="0563C1",
                    underline="single",
                )

    # Apply a compact default row height.
    for row_number in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 32

    worksheet.sheet_view.showGridLines = False

    workbook.save(excel_path)


def export_role_target_contacts(role_targets: list[dict]) -> dict:
    """
    Exports role-target contacts to CSV and formatted Excel files.

    Includes primary-source information and any additional sources preserved
    during cross-source contact merging.

    Returns file paths and export count.
    """
    ensure_output_dir()

    timestamp = get_timestamp()

    csv_path = OUTPUT_DIR / f"role_target_contacts_{timestamp}.csv"
    excel_path = OUTPUT_DIR / f"role_target_contacts_{timestamp}.xlsx"

    columns = [
        "organization",
        "name",
        "title",
        "department",
        "phone",
        "email",
        "role_categories",
        "matched_role_terms",
        "role_priority_score",
        "source_type",
        "source_page_title",
        "source_url",
        "additional_source_types",
        "additional_source_page_titles",
        "additional_source_urls",
    ]

    rows = []

    for contact in role_targets:
        row = {}

        for column in columns:
            value = contact.get(column, "")

            if isinstance(value, list):
                value = " | ".join(str(item) for item in value)

            row[column] = value

        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    df.to_csv(
        csv_path,
        index=False,
        encoding="utf-8-sig",
    )

    df.to_excel(
        excel_path,
        index=False,
    )

    format_role_target_excel(excel_path)

    return {
        "count": len(rows),
        "csv_path": str(csv_path),
        "excel_path": str(excel_path),
    }